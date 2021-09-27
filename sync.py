import argparse
import openpyxl
import sys
import glob
import os
from pathlib import Path
from proknow import ProKnow
from tqdm import tqdm


########################################
# Command Line Arguments

# Define arguments
parser = argparse.ArgumentParser(description="Synchronize ProKnow identity and access management.",
    formatter_class=argparse.ArgumentDefaultsHelpFormatter)
parser.add_argument("-u", "--url",
    required=True, help="the base URL to use when making request to the ProKnow API")
parser.add_argument("-c", "--credentials",
    required=True, help='path to ProKnow API credentials file')
parser.add_argument("--workspaces-file", default="workspaces.xlsx", 
    help="name of Excel workbook containing desired workspaces (within data directory)")
parser.add_argument("--workspace-slug-column", default="Slug",
    help="column name in workspaces Excel workbook that contains the workspace 'slug' field")
parser.add_argument('--workspace-name-column', default="Name",
    help="column name in workspaces Excel workbook that contains the workspace 'name' field")
parser.add_argument("--roles-file", default="roles.xlsx", 
    help="name of Excel workbook containing desired roles (within data directory)")
parser.add_argument("--users-directory", default="users", 
    help="name of directory containing user Excel workbooks (within data directory)")
parser.add_argument("--user-workspace-column", default="Workspace",
    help="column name in users Excel workbook that contains the 'slug' of the primary workspace")
parser.add_argument('--user-name-column', default="Name",
    help="column name in users Excel workbook that contains the user 'name' field")
parser.add_argument('--user-email-column', default="Email",
    help="column name in users Excel workbook that contains the user 'email' field")
parser.add_argument('--user-role-column', default="Role",
    help="column name in users Excel workbook that contains the name of the desired role")
parser.add_argument('--user-active-column', default="Active",
    help="column name in users Excel workbook that contains the user 'active' field")
parser.add_argument("data", help="directory containing workspace, role, and user records")

# Parse arguments
args = parser.parse_args()


########################################
# Logging

def beep(): sys.stdout.write("\a")

def print_blue(skk): print("\033[94m{}\033[00m".format(skk))

def print_cyan(skk): print("\033[96m{}\033[00m".format(skk))

def print_green(skk): print("\033[92m{}\033[00m".format(skk))

def print_magenta(skk): print("\033[95m{}\033[00m".format(skk))

def print_red(skk): print("\033[91m{}\033[00m".format(skk))

def print_yellow(skk): print("\033[93m{}\033[00m".format(skk))

def fail(skk, msg=None):
    print_red(skk)
    if msg:
        print_yellow(msg)
    beep()
    sys.exit(1)


########################################
# Utilities

def parse_bool(value):
    if type(value) == bool:
        return value
    elif type(value) == str:
        x = value.lower()
        return x == 'true' or x == 'yes'
    else:
        return None

def set_prop(obj, path, value):
    keys = path.split('.')
    cur = obj
    index = 0
    while index < len(keys) - 1:
        key = keys[index]
        if key not in obj:
            obj[key] = {}
        cur = obj[key]
        index += 1
    cur[keys[index]] = value

def confirm(question, default="yes"):
    valid = { "yes": True, "y": True, "ye": True, "no": False, "n": False }
    if default is None:
        prompt = " [y/n] "
    elif default == "yes":
        prompt = " [Y/n] "
    elif default == "no":
        prompt = " [y/N] "
    else:
        raise ValueError("invalid default answer: '%s'" % default)
    while True:
        sys.stdout.write(question + prompt)
        choice = input().lower()
        if default is not None and choice == "":
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' " "(or 'y' or 'n').\n")

def resolve_headers(filename, ws, options):
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for index, value in enumerate(row):
            if type(value) is str:
                v = value.strip().lower()
                for key, head in options.items():
                    if v == head:
                        if key in headers:
                            fail("Failed to resolve headers in '{0}' workbook".format(filename), (
                                "Duplicate '" + v + "' columns"
                            ))
                        headers[key] = index
                        break
    for key, head in options.items():
        if key not in headers:
            fail("Failed to resolve headers in '{0}' workbook".format(filename), (
                "Missing '" + key + "' column"
            ))
    return headers


########################################
# ProKnow Client

pk = ProKnow(args.url, credentials_file=args.credentials)


########################################
# Workspaces

print_magenta("Synchronizing Workspaces...")

# Load workspaces workbook
wb = openpyxl.load_workbook(Path(args.data, args.workspaces_file))
if "Workspaces" not in wb.sheetnames:
    fail("Failed to read '{0}' workspace workbook".format(Path(args.data, args.workspaces_file)), (
        "Workbook must contain 'Workspaces' sheet"
    ))
ws = wb["Workspaces"]

# Resolve headers
headers = resolve_headers(args.workspaces_file, ws, {
    "slug": args.workspace_slug_column.strip().lower(),
    "name": args.workspace_name_column.strip().lower()
})

# Read desired workspaces from workbook
workspaces = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[headers["slug"]] and row[headers["name"]]:
        slug = row[headers["slug"]].strip().lower()
        name = "[{0}] {1}".format(slug.upper(), row[headers["name"]].strip())
        workspaces[slug] = {
            "slug": slug,
            "name": name,
            "item": None
        }

# Query for existing workspaces and associate with defined workspaces
unknown_workspaces = []
for workspace_item in pk.workspaces.query():
    if workspace_item.slug in workspaces:
        workspaces[workspace_item.slug]["item"] = workspace_item
    else:
        unknown_workspaces.append(workspace_item)

# Identify created and updated workspaces
created = 0
updated = 0
workspace_jobs = []
for slug, workspace in workspaces.items():
    if workspace["item"] is None:
        workspace_jobs.append(workspace)
        created += 1
    elif workspace["item"].name != workspace["name"]:
        workspace["item"].name = workspace["name"]
        workspace_jobs.append(workspace)
        updated += 1

# Confirm creation and updating of workspaces
if len(workspace_jobs) > 0:
    print_yellow(" Workspaces have changed ({0} created, {1} updated)".format(created, updated))
    if not confirm("Are you sure you wish to synchronize workspaces?"):
        fail("Synchronization aborted")
    for workspace in tqdm(workspace_jobs):
        if workspace["item"] is None:
            workspace["item"] = pk.workspaces.create(workspace["slug"], workspace["name"])
        else:
            workspace["item"].save()
    print_green(" Workspaces successfully synchronized")
else:
    print_green(" All {0} workspaces are up to date".format(len(workspaces.keys())))


########################################
# Role Templates

print_magenta("Reading Role Templates...")

# Define static role permission lookup
role_permission_lookup = {
    "Advanced User Permissions": {
        "Create API Keys": "organization.create_api_keys"
    },
    "Organization Management Permissions": {
        "Manage Users, Roles, and Workspaces": "organization.manage_access",
        "Manage Custom Metrics": "organization.manage_custom_metrics",
        "Manage Renaming Rules": "organization.manage_template_metric_sets",
        "Manage Scorecard Templates": "organization.manage_renaming_rules",
        "Manage Checklist Templates": "organization.manage_template_checklists",
        "Manage Structure Set Templates": "organization.manage_template_structure_sets",
        "Manage Workspace Algorithms": "organization.manage_workspace_algorithms"
    },
    "All Workspaces": {
        "Read Patients": "organization.organization_read_patients",
        "Manage Patient Access": "organization.organization_manage_access_patients",
        "View PHI": "organization.organization_view_phi",
        "Download DICOM": "organization.organization_download_dicom",
        "Upload DICOM": "organization.organization_upload_dicom",
        "Write Patients": "organization.organization_write_patients",
        "Contour Patients": "organization.organization_contour_patients",
        "Delete Patients": "organization.organization_delete_patients",
        "Read Collections": "organization.organization_read_collections",
        "Write Collections": "organization.organization_write_collections",
        "Delete Collections": "organization.organization_delete_collections",
        "Collaborator": "organization.organization_collaborator"
    },
    "Primary Workspaces": {
        "Read Patients": "primary_workspaces.read_patients",
        "Manage Patient Access": "primary_workspaces.manage_access_patients",
        "View PHI": "primary_workspaces.view_phi",
        "Download DICOM": "primary_workspaces.download_dicom",
        "Upload DICOM": "primary_workspaces.upload_dicom",
        "Write Patients": "primary_workspaces.write_patients",
        "Contour Patients": "primary_workspaces.contour_patients",
        "Delete Patients": "primary_workspaces.delete_patients",
        "Read Collections": "primary_workspaces.read_collections",
        "Write Collections": "primary_workspaces.write_collections",
        "Delete Collections": "primary_workspaces.delete_collections",
        "Collaborator": "primary_workspaces.collaborator"
    },
    "Other Workspaces": {
        "Read Patients": "other_workspaces.read_patients",
        "Manage Patient Access": "other_workspaces.manage_access_patients",
        "View PHI": "other_workspaces.view_phi",
        "Download DICOM": "other_workspaces.download_dicom",
        "Upload DICOM": "other_workspaces.upload_dicom",
        "Write Patients": "other_workspaces.write_patients",
        "Contour Patients": "other_workspaces.contour_patients",
        "Delete Patients": "other_workspaces.delete_patients",
        "Read Collections": "other_workspaces.read_collections",
        "Write Collections": "other_workspaces.write_collections",
        "Delete Collections": "other_workspaces.delete_collections",
        "Collaborator": "other_workspaces.collaborator"
    }
}

# Load roles workbook and read desired role templates (from individual worksheets)
wb = openpyxl.load_workbook(Path(args.data, args.roles_file))
role_templates = {}
for sheetname in wb.sheetnames:

    # Initialize role template
    role_template = {
        "name": None
    }
    for category, permission_set in role_permission_lookup.items():
        for permission_name, permission_id in permission_set.items():
            set_prop(role_template, permission_id, False)

    # Parse permissions from worksheet rows
    ws = wb[sheetname]
    category = None
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] == 'Name':
            if row[1]:
                role_template['name'] = row[1]
            else:
                fail("Failed to read '{0}' role template definition".format(sheetname), (
                    "Invalid 'Name' specification, value must be provided"
                ))
        elif row[0] and row[1] is None:
            category = row[0].strip()
            if category not in role_permission_lookup:
                fail("Failed to read '{0}' role template definition".format(sheetname), (
                    "Invalid permission category '" + category + "'"
                ))
        elif row[0] and row[1]:
            if category:
                if row[0] in role_permission_lookup[category]:
                    value = row[1].lower() == 'yes'
                    set_prop(role_template, role_permission_lookup[category][row[0]], value)
                else:
                    fail("Failed to read '{0}' role template definition".format(sheetname), (
                        "Invalid permission '" + row[0] + "' (value '" + row[1] + "') "
                        "in category '" + category + "'"
                    ))
            else:
                fail("Failed to read '{0}' role template definition".format(sheetname), (
                    "Invalid attempt to specify permission '" + row[0] +
                    "' (value '" + row[1] + "') outside of a category"
                ))
        elif row[0] or row[1]:
            fail("Invalid role template definition in '{0}' sheet".format(sheetname), (
                "Invalid row contents (" + 
                ("'" + row[0] + "'" if isinstance(row[0], str) else 'None') + ", " + 
                ("'" + row[1] + "'" if isinstance(row[1], str) else 'None') + ")"
            ))

    # Add to role template
    if role_template["name"]:
        name = role_template["name"]
        if name != sheetname:
            fail("Invalid role template definition in '{0}' sheet".format(sheetname), (
                "Sheet name does not match role template name '" + name + "' specified"
            ))            
        if name not in role_templates:
            role_templates[name] = role_template
        else:
            fail("Invalid role template definition in '{0}' sheet".format(sheetname), (
                "Role template with name '" + name + "' already defined"
            ))
    else:
        fail("Invalid role template definition in '{0}' sheet".format(sheetname), (
            "Role template name must be specified"
        ))

# Report results
print_green(" Found {0} role templates".format(len(role_templates.keys())))


########################################
# Users

print_magenta("Reading Users...")

# Read users from worksheets
users = {}
user_files = glob.glob(str(Path(args.data, args.users_directory, "[!~$]*.xlsx")))
for user_file in user_files:
    wb = openpyxl.load_workbook(Path(user_file))
    if "Users" not in wb.sheetnames:
        fail("Failed to parse users from '{0}'".format(user_file), (
            "Workbook must contain 'Users' sheet"
        ))
    ws = wb['Users']

    # Resolve headers
    headers = resolve_headers(user_file, ws, {
        "workspace": args.user_workspace_column.strip().lower(),
        "name": args.user_name_column.strip().lower(),
        "email": args.user_email_column.strip().lower(),
        "role": args.user_role_column.strip().lower(),
        "active": args.user_active_column.strip().lower(),
    })

    # Parse users
    user_row = {}
    for row in ws.iter_rows(min_row=2):

        # Parse column values
        row_empty = True
        row_none = False
        for col, index in headers.items():
            if row[index].value is not None:
                val = row[index].value
                if col == 'email' or col == 'workspace':
                    user_row[col] = val.strip().lower()
                elif col == 'active':
                    user_row[col] = parse_bool(val)
                else:
                    user_row[col] = val.strip()
                row_empty = False
            else:
                row_none = col

        # Validate and record user data
        if not row_empty:
            if row_none:
                fail("Failed to parse users from '{0}'".format(user_file), (
                    "User is missing '" + row_none.title() + "' value in row " + str(row[0].row)
                ))
            else:

                # Validate workspace
                if user_row["workspace"] not in workspaces:
                    fail("Failed to parse users from '{0}'".format(user_file), (
                        "User at row " + str(row[0].row) + " references an unknown workspace "
                        "'" + user_row["workspace"] + "'"
                    ))

                # Create user object (if it does not already exist)
                if user_row["email"] not in users:
                    users[user_row["email"]] = {
                        "name": user_row["name"],
                        "email": user_row["email"],
                        "active": user_row["active"],
                        "workspaces": {},
                        "item": None,
                        "role": None
                    }
                
                # Validate user and append workspace role
                user = users[user_row["email"]]
                if user_row["workspace"] in user["workspaces"]:
                    fail("Failed to parse users from '{0}'".format(user_file), (
                        "User at row " + str(row[0].row) + " has multiple role assignments "
                        "for workspace '" + user_row["workspace"] + "'" + os.linesep + "First " 
                        "assigned in '" + user["workspaces"][user_row["workspace"]]["file"] + "'"
                    ))
                for slug, workspace in user["workspaces"].items():
                    if workspace["role"] != user_row["role"]:
                        fail("Failed to parse users from '{0}'".format(user_file), (
                            "User at row " + str(row[0].row) + " has conflicting role assignments "
                            "of role '" + user_row["role"] + "' and '" + workspace["role"] + "'" +
                            os.linesep + "First assigned in '" + workspace["file"] + "'"
                        ))

                # Add user to collection
                user["workspaces"][user_row["workspace"]] = {
                    "role": user_row["role"],
                    "workspace": user_row["workspace"],
                    "file": user_file
                }

# Report results
print_green(" Found {0} users".format(len(users.keys())))


########################################
# Roles

print_magenta("Synchronizing Roles...")

# Create required roles
roles = {}
for email, user in users.items():

    # Determine primary workspaces for role
    role_template_name = None
    role_workspaces = []
    for slug, workspace in user["workspaces"].items():
        role_template_name = workspace["role"]
        role_workspaces.append(slug)

    # Create role
    role_name = "[" + "+".join(role_workspaces).upper() + "] " + role_template_name
    if role_name not in roles:
        role_template = role_templates[role_template_name]
        role = {
            "name": role_name,
            "item": None,
            "data": {
                "workspaces": []
            }
        }

        # Set organization permissions
        for key, value in role_template["organization"].items():
            set_prop(role["data"], key, value)

        # Set primary workspace permissions
        for slug in role_workspaces:
            if slug not in workspaces or workspaces[slug]["item"] is None:
                fail("Failed to create role '{0}'".format(role_name), (
                    "Workspace '" + slug + "' not found"
                ))
            wp = {
                "id": workspaces[slug]["item"].id
            }
            for key, value in role_template["primary_workspaces"].items():
                set_prop(wp, key, value)
            role["data"]["workspaces"].append(wp)

        # Set other workspace permissions
        other_workspaces_required = False
        for key, value in role_template["other_workspaces"].items():
            if value is True:
                other_workspaces_required = True
                break
        if other_workspaces_required:
            for slug, workspace in workspaces.items():
                if slug not in role_workspaces:
                    wp = {
                        "id": workspace["item"].id
                    }
                    for key, value in role_template["other_workspaces"].items():
                        set_prop(wp, key, value)
                    role["data"]["workspaces"].append(wp)

        # Sort workspaces and add role to collection
        role["data"]["workspaces"].sort(key=lambda ws: ws["id"])
        roles[role_name] = role

    # Associate role with user
    user["role"] = roles[role_name]

# Query for existing roles and associate
unknown_roles = []
for role_item in [role.get() for role in pk.roles.query()]:
    role_item.permissions["workspaces"].sort(key=lambda ws: ws["id"])
    if role_item.name in roles:
        role_item.permissions.pop('private', None)
        role_item.permissions.pop('user', None)
        roles[role_item.name]["item"] = role_item
    elif role_item.name != 'Admin':
        unknown_roles.append(role_item)

# Identify created and updated workspaces
created = 0
updated = 0
role_jobs = []
for name, role in roles.items():
    if role["item"] is None:
        role_jobs.append(role)
        created += 1
    elif role["data"] != role["item"].permissions:
        role["item"].permissions = role["data"]
        role_jobs.append(role)
        updated += 1

# Confirm creation and updating of roles
if len(role_jobs) > 0:
    print_yellow(" Roles have changed ({0} created, {1} updated)".format(created, updated))
    if not confirm("Are you sure you wish to synchronize roles?"):
        fail("Synchronization aborted")
    for role in tqdm(role_jobs):
        if role["item"] is None:
            role["item"] = pk.roles.create(role["name"], role["data"])
        else:
            role["item"].permissions["private"] = False
            role["item"].permissions["user"] = None
            role["item"].save()
    print_green(" Roles successfully synchronized")
else:
    print_green(" All {0} roles are up to date".format(len(roles.keys())))


########################################
# Users

print_magenta("Synchronizing Users...")

# Query for existing users and associate with users defined in spreadsheets
unknown_users = []
for user_item in pk.users.query():
    if user_item.email in users:
        users[user_item.email]["item"] = user_item
    else:
        unknown_users.append(user_item)

# Identify created and updated workspaces
created = 0
updated = 0
user_jobs = []
for email, user in users.items():
    item = user["item"]
    if item is None:
        user_jobs.append(user)
        created += 1
    elif (user["name"] != item.name or user["active"] != item.data["active"] or 
            user["role"]["item"].id != item.data["role"]["id"]):
        user_jobs.append(user)
        updated += 1

# Confirm creation and updating of users
if len(user_jobs) > 0:
    print_yellow(" Users have changed ({0} created, {1} updated)".format(created, updated))
    if not confirm("Are you sure you wish to synchronize users?"):
        fail("Synchronization aborted")
    for user in tqdm(user_jobs):
        if user["item"] is None:
            user["item"] = pk.users.create(user["email"], user["name"], user["role"]["item"].id)
        else:
            item = user["item"].get()
            item.name = user["name"]
            item.active = user["active"]
            item.role_id = user["role"]["item"].id
            item.save()
    print_green(" Users successfully synchronized")
else:
    print_green(" All {0} users are up to date".format(len(users.keys())))


########################################
# Reporting

if len(unknown_workspaces) > 0 or len(unknown_roles) > 0 or len(unknown_users) > 0:
    print_magenta("Identifying Unknown Resources...")
    if len(unknown_workspaces) > 0:
        print_yellow(" Identified {0} unknown workspaces:".format(len(unknown_workspaces)))
        for workspace in unknown_workspaces:
            print("  {0} ({1})".format(workspace.name, workspace.slug))
    if len(unknown_roles) > 0:
        print_yellow(" Identified {0} unknown roles:".format(len(unknown_roles)))
        for role in unknown_roles:
            print("  {0}".format(role.name))
    if len(unknown_users) > 0:
        print_yellow(" Identified {0} unknown users:".format(len(unknown_users)))
        for user in unknown_users:
            print("  {0} ({1})".format(user.name, user.email))
